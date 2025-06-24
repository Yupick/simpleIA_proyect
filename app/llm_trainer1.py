#!/usr/bin/env python3
"""
llm_trainer.py
--------------
Script de entrenamiento para el LLM que permite:
  - Seleccionar un modelo preentrenado.
  - Realizar fine tuning con datos disponibles en las carpetas 'dialogue' y 'knowledge'.
  - Entrenamiento manual y reentrenamiento con feedback.
  
Estructura (relativa a la raíz del proyecto):
  - config/config.json
  - trainer_llm/dialogue y trainer_llm/knowledge para datos de entrenamiento.
  - model_llm para almacenar modelos entrenados.
  - feedback/feedback.sqlite para almacenar feedback.
"""

import os
import json
import time
import sqlite3
import shutil
from pathlib import Path

import pandas as pd
import fitz          # PyMuPDF, para PDF
import docx
from openpyxl import load_workbook
from transformers import AutoTokenizer, AutoModelForCausalLM, Trainer, TrainingArguments
from datasets import Dataset
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# La ruta base se asume que es el directorio raíz del proyecto.
# Dado que este script está en "app/", la raíz será el directorio padre.
BASE_DIR = Path(__file__).resolve().parent.parent

CONFIG_PATH = BASE_DIR / "config" / "config.json"
TRAINING_DIR = BASE_DIR / "trainer_llm"
DIALOGUE_DIR = TRAINING_DIR / "dialogue"
KNOWLEDGE_DIR = TRAINING_DIR / "knowledge"
MODEL_DIR = BASE_DIR / "model_llm"
FEEDBACK_DB = BASE_DIR / "feedback" / "feedback.sqlite"
SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".csv", ".json", ".xlsx", ".xls", ".docx"}

# Función para cargar la configuración
def load_config():
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                config = json.load(f)
                logger.info("Config loaded.")
                return config
        except Exception as e:
            logger.error(f"Error loading config: {e}")
            return {}
    else:
        logger.info("config.json not found. Using empty config.")
        return {}

# Función para guardar la configuración
def save_config(config):
    try:
        CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4)
        logger.info("Config saved in config.json.")
    except Exception as e:
        logger.error(f"Error saving config: {e}")

# Función para leer diferentes tipos de archivo y extraer líneas de texto
def parse_file(file_path: Path):
    ext = file_path.suffix.lower()
    try:
        if ext == ".txt":
            return file_path.read_text(encoding="utf-8").splitlines()
        elif ext == ".pdf":
            text = ""
            with fitz.open(str(file_path)) as doc:
                for page in doc:
                    text += page.get_text()
            return text.splitlines()
        elif ext == ".csv":
            df = pd.read_csv(file_path)
            return df.astype(str).apply(lambda row: " ".join(row), axis=1).tolist()
        elif ext == ".json":
            df = pd.read_json(file_path)
            return df.astype(str).apply(lambda row: " ".join(row), axis=1).tolist()
        elif ext in [".xlsx", ".xls"]:
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            lines = []
            for row in sheet.iter_rows():
                line = " ".join(str(cell.value) for cell in row if cell.value is not None)
                if line:
                    lines.append(line)
            return lines
        elif ext == ".docx":
            doc = docx.Document(file_path)
            return [para.text for para in doc.paragraphs if para.text.strip()]
        else:
            logger.warning(f"Unsupported extension: {file_path.name}")
            return []
    except Exception as e:
        logger.error(f"Error reading {file_path.name}: {e}")
        return []

# Función para analizar un directorio y extraer datos de los archivos válidos
def collect_training_data(directory: Path):
    logger.info(f"Analyzing folder: {directory.name}")
    texts = []
    invalid_files = []
    file_count = 0
    line_count = 0
    if not directory.exists():
        logger.warning(f"Directory '{directory}' not found.")
        return texts, file_count, line_count
    for file in directory.glob("*"):
        if file.suffix.lower() in SUPPORTED_EXTENSIONS:
            lines = parse_file(file)
            if lines:
                file_count += 1
                line_count += len(lines)
                texts.extend(lines)
            else:
                invalid_files.append(file)
        else:
            invalid_files.append(file)
    logger.info(f"Valid files: {file_count}, lines extracted: {line_count}")
    if invalid_files:
        logger.warning("Invalid files:")
        for f in invalid_files:
            logger.warning(f" - {f.name}")
    return texts, file_count, line_count

# Función para cargar el modelo preentrenado sin fine tuning.
def load_model():
    """
    Carga el modelo preentrenado basado en la configuración.
    Se utiliza cuando no se realiza fine tuning.
    """
    config = load_config()
    model_name = config.get("selected_model", "gpt2")
    logger.info(f"Loading pretrained model: {model_name}")
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForCausalLM.from_pretrained(model_name)
        logger.info("Pretrained model loaded successfully.")
    except Exception as e:
        logger.error(f"Error loading model {model_name}: {e}")

# Función para entrenar el modelo a partir de los datos
def train_model(model_name: str, training_data: list):
    if not training_data:
        logger.error("Not enough data to train.")
        return
    try:
        logger.info(f"Starting training with model {model_name}...")
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForCausalLM.from_pretrained(model_name)
    except Exception as e:
        logger.error(f"Error loading model {model_name}: {e}")
        return

    def tokenize_function(example):
        return tokenizer(example["text"], padding="max_length", truncation=True, max_length=128)
    try:
        logger.info("Preparing dataset...")
        raw_dataset = Dataset.from_dict({"text": training_data})
        tokenized = raw_dataset.map(tokenize_function, batched=True)
    except Exception as e:
        logger.error(f"Error preparing dataset: {e}")
        return
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    output_path = MODEL_DIR / f"{model_name.replace('/', '_')}_{timestamp}"
    try:
        output_path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.error(f"Error creating model directory: {e}")
        return
    config_vals = load_config()
    num_train_epochs = config_vals.get("num_train_epochs", 3)
    per_device_train_batch_size = config_vals.get("per_device_train_batch_size", 4)
    training_args = TrainingArguments(
        output_dir=str(output_path),
        overwrite_output_dir=True,
        num_train_epochs=num_train_epochs,
        per_device_train_batch_size=per_device_train_batch_size,
        save_steps=10,
        save_total_limit=2,
        logging_dir=str(output_path / "logs"),
        logging_steps=5,
        evaluation_strategy="no",
        report_to=[]
    )
    try:
        logger.info("Training the model...")
        trainer = Trainer(
            model=model,
            args=training_args,
            train_dataset=tokenized,
            tokenizer=tokenizer
        )
        trainer.train()
        trainer.save_model(output_path)
        tokenizer.save_pretrained(output_path)
        logger.info(f"Training complete. Model saved at: {output_path}")
    except Exception as e:
        logger.error(f"Training error: {e}")

# Función para entrenamiento manual ingresando datos por consola
def train_manual_input():
    logger.info("Manual training mode started.")
    data = []
    logger.info("Enter text lines (empty line to finish):")
    while True:
        line = input("> ").strip()
        if not line:
            break
        data.append(line)
    if data:
        model_name = load_config().get("selected_model", "gpt2")
        train_model(model_name, data)
    else:
        logger.warning("No data entered for manual training.")

# Función para reentrenar usando el feedback almacenado
def retrain_with_feedback():
    if not FEEDBACK_DB.exists():
        logger.error("Feedback database not found.")
        return
    try:
        conn = sqlite3.connect(str(FEEDBACK_DB))
        cursor = conn.cursor()
        cursor.execute("SELECT text FROM feedback;")
        rows = cursor.fetchall()
        logger.info(f"Feedback available: {len(rows)} lines")
        if not rows:
            return
        confirm = input("Retrain with feedback? (y/n): ").strip().lower()
        if confirm == "y":
            data = [r[0] for r in rows]
            model_name = load_config().get("selected_model", "gpt2")
            train_model(model_name, data)
        else:
            logger.info("Retraining canceled.")
    except Exception as e:
        logger.error(f"Feedback error: {e}")
    finally:
        conn.close()

# Función para eliminar el modelo entrenado previamente
def delete_trained_model():
    if MODEL_DIR.exists():
        confirm = input("Delete trained model and all its data? (y/n): ").strip().lower()
        if confirm == "y":
            try:
                shutil.rmtree(MODEL_DIR)
                logger.info("Trained model deleted successfully.")
            except Exception as e:
                logger.error(f"Error deleting model: {e}")
    else:
        logger.info("No trained model to delete.")

# Función para seleccionar un modelo preentrenado y, opcionalmente, entrenarlo
def select_pretrained_model():
    english_models = ["gpt2", "bigscience/bloom-560m", "EleutherAI/gpt-neo-125M"]
    spanish_models = ["datificate/gpt2-small-spanish", "flax-community/gpt-2-spanish"]
    print("\n📦 Available Models:")
    print("\nEnglish Models:")
    for idx, model in enumerate(english_models, 1):
        print(f"  {idx}. {model}")
    offset = len(english_models)
    print("\nSpanish Models:")
    for idx, model in enumerate(spanish_models, offset + 1):
        print(f"  {idx}. {model}")
    print("  0. Back")
    choice = input("Select a model: ").strip()
    if choice == "0":
        return
    try:
        choice = int(choice)
        total_models = len(english_models) + len(spanish_models)
        if choice < 1 or choice > total_models:
            print("Invalid option.")
            return
        if choice <= len(english_models):
            selected = english_models[choice - 1]
        else:
            selected = spanish_models[choice - len(english_models)]
        print(f"\nSelected Model: {selected}")
        config = load_config()
        config["selected_model"] = selected
        save_config(config)
        choice_training = input("Perform fine tuning with available data? (y/n): ").strip().lower()
        if choice_training == "y":
            dialogue_data, _, _ = collect_training_data(DIALOGUE_DIR)
            knowledge_data, _, _ = collect_training_data(KNOWLEDGE_DIR)
            training_data = dialogue_data + knowledge_data
            if not training_data:
                print("Not enough data for training, loading pretrained model without training.")
                load_model()
            else:
                train_model(selected, training_data)
        else:
            load_model()
    except (ValueError, IndexError):
        print("Invalid option.")

# Menú principal
def main_menu():
    while True:
        print("\n=== MAIN MENU ===")
        print("1. Select pretrained model")
        print("2. Delete trained model")
        print("3. Train with files (dialogue/knowledge)")
        print("4. Manual training")
        print("5. Retrain with user feedback")
        print("0. Exit")
        choice = input("Select an option: ").strip()
        if choice == "1":
            select_pretrained_model()
        elif choice == "2":
            delete_trained_model()
        elif choice == "3":
            dialogue_data, _, _ = collect_training_data(DIALOGUE_DIR)
            knowledge_data, _, _ = collect_training_data(KNOWLEDGE_DIR)
            total_data = dialogue_data + knowledge_data
            print(f"\nTotal lines for training: {len(total_data)}")
            if not total_data:
                logger.error("No data found for training.")
            else:
                confirm = input("Continue training? (y/n): ").strip().lower()
                if confirm == "y":
                    model_name = load_config().get("selected_model", "gpt2")
                    train_model(model_name, total_data)
        elif choice == "4":
            train_manual_input()
        elif choice == "5":
            retrain_with_feedback()
        elif choice == "0":
            print("Exiting.")
            break
        else:
            print("Invalid option, try again.")

if __name__ == "__main__":
    main_menu()
