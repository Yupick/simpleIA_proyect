"""Trainer unificado (remplaza legacy llm_trainer.py).
Funcionalidades:
  - Seleccionar modelo preentrenado.
  - Cargar datos desde carpetas dialogue/knowledge.
  - Entrenamiento manual.
  - Re-entrenamiento con feedback.
"""
import shutil
import time
import logging
import sqlite3
from pathlib import Path
from typing import List
import pandas as pd
import fitz
import docx
from openpyxl import load_workbook
from datasets import Dataset
from transformers import (
    AutoTokenizer,
    AutoModelForCausalLM,
    Trainer,
    TrainingArguments,
    DataCollatorForLanguageModeling,
)
from ..core.config import config

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent.parent
TRAINING_DIR = BASE_DIR / "trainer_llm"
DIALOGUE_DIR = TRAINING_DIR / "dialogue"
KNOWLEDGE_DIR = TRAINING_DIR / "knowledge"
MODEL_DIR = BASE_DIR / "model_llm"
FEEDBACK_DB = BASE_DIR / "feedback" / "feedback.sqlite"
SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".csv", ".json", ".xlsx", ".xls", ".docx"}


def _parse_file(path: Path) -> List[str]:
    ext = path.suffix.lower()
    try:
        if ext == ".txt":
            return path.read_text(encoding="utf-8").splitlines()
        if ext == ".pdf":
            text = ""
            with fitz.open(str(path)) as doc:
                for page in doc:
                    text += page.get_text()
            return text.splitlines()
        if ext == ".csv":
            df = pd.read_csv(path)
            return df.astype(str).apply(lambda r: " ".join(r), axis=1).tolist()
        if ext == ".json":
            df = pd.read_json(path)
            return df.astype(str).apply(lambda r: " ".join(r), axis=1).tolist()
        if ext in [".xlsx", ".xls"]:
            wb = load_workbook(path, read_only=True)
            sheet = wb.active
            lines = []
            for row in sheet.iter_rows():
                line = " ".join(str(c.value) for c in row if c.value is not None)
                if line:
                    lines.append(line)
            return lines
        if ext == ".docx":
            docx_doc = docx.Document(path)
            return [p.text for p in docx_doc.paragraphs if p.text.strip()]
        logger.warning(f"Unsupported extension: {path.name}")
        return []
    except Exception as e:
        logger.error(f"Error reading {path.name}: {e}")
        return []


def _collect(directory: Path) -> List[str]:
    if not directory.exists():
        logger.warning(f"Directory missing: {directory}")
        return []
    data: List[str] = []
    for f in directory.glob("*"):
        if f.suffix.lower() in SUPPORTED_EXTENSIONS:
            lines = _parse_file(f)
            data.extend(lines)
    logger.info(f"Collected {len(data)} lines from {directory.name}")
    return data


def _prepare_dataset(lines: List[str], model_name: str):
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForCausalLM.from_pretrained(model_name)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    def tokenize(example):
        t = tokenizer(example["text"], padding="max_length", truncation=True, max_length=128)
        t["labels"] = t["input_ids"].copy()
        return t

    raw = Dataset.from_dict({"text": lines})
    tokenized = raw.map(tokenize, batched=True)
    return tokenizer, model, tokenized


def train_files():
    model_name = config.selected_model
    dialogue = _collect(DIALOGUE_DIR)
    knowledge = _collect(KNOWLEDGE_DIR)
    lines = dialogue + knowledge
    if not lines:
        logger.error("No hay datos para entrenar.")
        return None
    return _train(model_name, lines)


def train_manual():
    logger.info("Entrenamiento manual: ingresa líneas (vacío para terminar)")
    lines = []
    while True:
        line = input("> ").strip()
        if not line:
            break
        lines.append(line)
    if not lines:
        logger.warning("Sin datos manuales.")
        return None
    return _train(config.selected_model, lines)


def train_feedback():
    if not FEEDBACK_DB.exists():
        logger.error("Base de feedback inexistente.")
        return None
    try:
        with sqlite3.connect(str(FEEDBACK_DB)) as conn:
            cur = conn.cursor()
            cur.execute("SELECT text FROM feedback")
            rows = cur.fetchall()
        lines = [r[0] for r in rows]
        if not lines:
            logger.warning("Sin feedback para entrenar.")
            return None
        confirm = input("Reentrenar con feedback? (y/n): ").strip().lower()
        if confirm != "y":
            logger.info("Cancelado.")
            return None
        return _train(config.selected_model, lines)
    except Exception as e:
        logger.error(f"Error leyendo feedback: {e}")
        return None


def delete_models():
    if not MODEL_DIR.exists():
        logger.info("No hay modelos para borrar.")
        return
    confirm = input("Eliminar carpeta de modelos completa? (y/n): ").strip().lower()
    if confirm == "y":
        shutil.rmtree(MODEL_DIR)
        logger.info("Modelos eliminados.")


def _train(model_name: str, lines: List[str]):
    logger.info(f"Inicio entrenamiento con {len(lines)} líneas usando modelo {model_name}")
    tokenizer, model, tokenized = _prepare_dataset(lines, model_name)
    ts = time.strftime("%Y%m%d-%H%M%S")
    out = MODEL_DIR / f"{model_name.replace('/', '_')}_{ts}"
    out.mkdir(parents=True, exist_ok=True)
    args = TrainingArguments(
        output_dir=str(out),
        overwrite_output_dir=True,
        num_train_epochs=config.num_train_epochs,
        per_device_train_batch_size=config.per_device_train_batch_size,
        save_steps=50,
        save_total_limit=2,
        logging_dir=str(out / "logs"),
        logging_steps=10,
    )
    collator = DataCollatorForLanguageModeling(tokenizer=tokenizer, mlm=False)
    trainer = Trainer(model=model, args=args, train_dataset=tokenized, data_collator=collator)
    trainer.train()
    trainer.save_model(out)
    tokenizer.save_pretrained(out)
    logger.info(f"Entrenamiento finalizado. Guardado en {out}")
    return str(out)


def main_menu():
    while True:
        print("\n=== TRAINER UNIFICADO ===")
        print("1. Entrenar con archivos (dialogue/knowledge)")
        print("2. Entrenamiento manual")
        print("3. Reentrenar con feedback")
        print("4. Borrar modelos entrenados")
        print("0. Salir")
        choice = input("Selecciona opción: ").strip()
        if choice == "1":
            train_files()
        elif choice == "2":
            train_manual()
        elif choice == "3":
            train_feedback()
        elif choice == "4":
            delete_models()
        elif choice == "0":
            break
        else:
            print("Opción inválida")


if __name__ == "__main__":
    main_menu()
