"""
Data Loader - Gestión de archivos de entrenamiento.
"""

import logging
from pathlib import Path
from typing import List, Dict, Tuple
import pandas as pd
import fitz  # PyMuPDF
import docx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".csv", ".json", ".xlsx", ".xls", ".docx"}


class TrainingDataLoader:
    """Carga y gestiona archivos de entrenamiento."""
    
    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        self.dialogue_dir = base_dir / "trainer_llm" / "dialogue"
        self.knowledge_dir = base_dir / "trainer_llm" / "knowledge"
        
        # Crear directorios si no existen
        self.dialogue_dir.mkdir(parents=True, exist_ok=True)
        self.knowledge_dir.mkdir(parents=True, exist_ok=True)
    
    def parse_file(self, file_path: Path) -> List[str]:
        """
        Lee un archivo y extrae líneas de texto según su extensión.
        
        Args:
            file_path: Ruta al archivo
            
        Returns:
            Lista de líneas de texto extraídas
        """
        ext = file_path.suffix.lower()
        try:
            if ext == ".txt":
                return file_path.read_text(encoding="utf-8").splitlines()
            
            elif ext == ".pdf":
                text = ""
                with fitz.open(str(file_path)) as doc:
                    for page in doc:
                        text += page.get_text()
                return text.splitlines()
            
            elif ext == ".csv":
                df = pd.read_csv(file_path)
                return df.astype(str).apply(lambda row: " ".join(row), axis=1).tolist()
            
            elif ext == ".json":
                df = pd.read_json(file_path)
                return df.astype(str).apply(lambda row: " ".join(row), axis=1).tolist()
            
            elif ext in [".xlsx", ".xls"]:
                wb = load_workbook(file_path, read_only=True)
                sheet = wb.active
                lines = []
                for row in sheet.iter_rows():
                    line = " ".join(str(cell.value) for cell in row if cell.value is not None)
                    if line:
                        lines.append(line)
                return lines
            
            elif ext == ".docx":
                doc = docx.Document(file_path)
                return [para.text for para in doc.paragraphs if para.text.strip()]
            
            else:
                logger.warning(f"Unsupported extension: {file_path.name}")
                return []
                
        except Exception as e:
            logger.error(f"Error reading {file_path.name}: {e}")
            return []
    
    def collect_from_directory(self, directory: Path) -> Tuple[List[str], int, int]:
        """
        Recopila datos de todos los archivos en un directorio.
        
        Args:
            directory: Directorio a analizar
            
        Returns:
            Tupla (textos, file_count, line_count)
        """
        logger.info(f"Analyzing folder: {directory.name}")
        texts = []
        file_count = 0
        line_count = 0
        
        if not directory.exists():
            logger.warning(f"Directory '{directory}' not found.")
            return texts, file_count, line_count
        
        for file in directory.glob("*"):
            if file.suffix.lower() in SUPPORTED_EXTENSIONS:
                lines = self.parse_file(file)
                if lines:
                    file_count += 1
                    line_count += len(lines)
                    texts.extend(lines)
        
        logger.info(f"Valid files: {file_count}, lines extracted: {line_count}")
        return texts, file_count, line_count
    
    def collect_all_data(self) -> Tuple[List[str], Dict[str, int]]:
        """
        Recopila todos los datos de ambas carpetas.
        
        Returns:
            Tupla (textos_combinados, estadísticas)
        """
        dialogue_data, dialogue_files, dialogue_lines = self.collect_from_directory(self.dialogue_dir)
        knowledge_data, knowledge_files, knowledge_lines = self.collect_from_directory(self.knowledge_dir)
        
        all_texts = dialogue_data + knowledge_data
        
        stats = {
            "dialogue_files": dialogue_files,
            "dialogue_lines": dialogue_lines,
            "knowledge_files": knowledge_files,
            "knowledge_lines": knowledge_lines,
            "total_lines": len(all_texts)
        }
        
        return all_texts, stats
    
    def collect_from_file(self, folder: str, filename: str) -> List[str]:
        """
        Recopila datos de un archivo específico.
        
        Args:
            folder: 'dialogue' o 'knowledge'
            filename: Nombre del archivo
            
        Returns:
            Lista de líneas de texto
        """
        directory = self.dialogue_dir if folder == "dialogue" else self.knowledge_dir
        file_path = directory / filename
        
        if not file_path.exists():
            raise FileNotFoundError(f"File {filename} not found in {folder}")
        
        return self.parse_file(file_path)
    
    def list_files(self, folder: str) -> List[Dict[str, any]]:
        """
        Lista archivos en una carpeta específica.
        
        Args:
            folder: 'dialogue' o 'knowledge'
            
        Returns:
            Lista de diccionarios con info de archivos
        """
        directory = self.dialogue_dir if folder == "dialogue" else self.knowledge_dir
        
        files = []
        for file in directory.glob("*"):
            if file.is_file() and file.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append({
                    "name": file.name,
                    "size": file.stat().st_size,
                    "extension": file.suffix,
                    "modified": file.stat().st_mtime
                })
        
        return sorted(files, key=lambda x: x["modified"], reverse=True)
    
    def delete_file(self, folder: str, filename: str) -> bool:
        """
        Elimina un archivo de una carpeta.
        
        Args:
            folder: 'dialogue' o 'knowledge'
            filename: Nombre del archivo a eliminar
            
        Returns:
            True si se eliminó exitosamente
        """
        directory = self.dialogue_dir if folder == "dialogue" else self.knowledge_dir
        file_path = directory / filename
        
        if not file_path.exists():
            raise FileNotFoundError(f"File {filename} not found in {folder}")
        
        try:
            file_path.unlink()
            logger.info(f"Deleted file: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error deleting file {filename}: {e}")
            raise
    
    def save_uploaded_file(self, folder: str, filename: str, content: bytes) -> Path:
        """
        Guarda un archivo subido en la carpeta correspondiente.
        
        Args:
            folder: 'dialogue' o 'knowledge'
            filename: Nombre del archivo
            content: Contenido binario del archivo
            
        Returns:
            Ruta del archivo guardado
        """
        directory = self.dialogue_dir if folder == "dialogue" else self.knowledge_dir
        file_path = directory / filename
        
        # Validar extensión
        if file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file extension: {file_path.suffix}")
        
        # Guardar archivo
        with open(file_path, "wb") as f:
            f.write(content)
        
        logger.info(f"Saved uploaded file: {file_path}")
        return file_path
